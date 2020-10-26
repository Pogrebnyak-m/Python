##

import openpyxl, random
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors

wb=Workbook()
ws=wb.active
ws.title="data"

x1 = ws.cell(row = 1, column = 1)
x1.value = "   Фамилия"
x2 = ws.cell(row = 1, column = 2)
x2.value = " Профвзносы"
x3 = ws.cell(row = 1, column = 3)
x3.value = " Месяц"
x4 = ws.cell(row = 1, column = 4)
x4.value = "Сумма"

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 21
ws.column_dimensions['C'].width = 13
ws.column_dimensions['D'].width = 11

ft_h = Font(name='Calibri Light',color='0033CCCC', size=20)
x1.font = ft_h
x2.font = ft_h
x3.font = ft_h
x4.font = ft_h

wb1=openpyxl.load_workbook('surnames.xlsx')
ws0=wb1.active
for i in range(0,30):
    ws['A'+str(i+2)]=ws0['A'+str(i+2)].value
month=['Все','Сентябрь','Февраль','Декабрь','Апрель']
for i in range(0,29):
    ws['C'+str(i+2)]=random.choice(month)

for i in range(0,29):
    ws['D'+str(i+2)]=random.randint(80,450)
for i in range(0,29):
    if (ws['D'+str(i+2)].value>170 and ws['C'+str(i+2)].value=='Все'):
        ws['B'+str(i+2)]='+'
    else: ws['B'+str(i+2)]='-'

ws2=wb.create_sheet(title="results")
y1 = ws2.cell(row = 1, column = 1)
y1.value = "   Фамилия"
y2 = ws2.cell(row = 1, column = 2)
y2.value = " Санаторий"
y3 = ws2.cell(row = 1, column = 3)
y3.value = "  Статус"
#y4 = ws2.cell(row = 1, column = 4)
#y4.value = " Количество"

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 15

ft_h = Font(name='Calibri Light',color='00003366', size=20)

y1.font = ft_h
y2.font = ft_h
y3.font = ft_h

wb2=openpyxl.load_workbook('surnames.xlsx')
for i in range(0,30):
    ws2['A'+str(i+2)]=ws0['A'+str(i+2)].value
sanatorium=['Волна','Лесной','Янтарь','Прибой','Искра']
for i in range(0,29):
    ws2['B'+str(i+2)]=random.choice(sanatorium)
for i in range(0,29):
    if (ws['B'+str(i+2)].value=='+'):
        ws2['C'+str(i+2)]='едет'
    else: ws2['C'+str(i+2)]='не едет'
    
#ws2['D2'] = "=СЧЁТЕСЛИ(C2:C30;"едет")"

wb.save('file.xlsx')
